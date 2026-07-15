# -*- coding: utf-8 -*-
"""merge_apply.py — apply one or more merge_*.json files onto the research workbook.
Usage: python3 merge_apply.py <workbook.xlsx> <merge.json> [<merge2.json> ...]
In-place update: Locations sheet (brand tags on matches, new rows appended), Research Status sheet.
"""
import sys, json, re, unicodedata
import openpyxl

WB = sys.argv[1]
merges = sys.argv[2:]
wb = openpyxl.load_workbook(WB)
ws = wb['Locations']

# locate header row & columns
hdr_row = None
for r in range(1, 15):
    if str(ws.cell(r, 2).value).strip() == '#':
        hdr_row = r; break
COL = {str(ws.cell(hdr_row, c).value).strip(): c for c in range(2, ws.max_column+1) if ws.cell(hdr_row,c).value}
C_NUM, C_GRP, C_MKT = COL['#'], COL['Group Name'], COL['Market']
C_CITY, C_ADDR, C_BR = COL['Location Name'], COL['Address'], COL['Brands Sold at Location']
C_TYPE = COL.get('Dealer Type')
C_VER = ws.max_column + 1
if str(ws.cell(hdr_row, C_VER-1).value).strip() != 'Tags verified':
    ws.cell(hdr_row, C_VER).value = 'Tags verified'
    ws.cell(hdr_row, C_VER+1).value = 'Verification source'
else:
    C_VER -= 1

# map DataFrame rid (pandas index from ingest) -> worksheet row: ingest used iloc offset hdr+1..; pandas index = excel row - 2? 
# Reconstruct mapping by re-reading with same logic:
import pandas as pd
loc = pd.read_excel(WB, sheet_name='Locations', header=None)
p_hdr = loc.index[loc[1].astype(str).str.strip() == '#'][0]
# pandas row i corresponds to excel row i+1
rid2xl = {int(i): int(i)+1 for i in loc.index}

def tagify(brands_str, brand, sales, service):
    """ensure brand appears with [S]/[SV] tag in the brands string"""
    tag = 'S' if sales else 'SV'
    if service and sales: tag = 'S,SV'
    parts = [p.strip() for p in re.split(r'[,;]', str(brands_str or '')) if p.strip() and p.strip().lower()!='nan']
    def base(p): return re.sub(r'\s*\[[^\]]*\]', '', re.sub(r'\(.*?\)', '', p)).strip().lower()
    hit = False
    for i, p in enumerate(parts):
        if base(p) == brand.lower() or brand.lower() in base(p):
            if '[' not in p: parts[i] = re.sub(r'\s*$', '', p) + f' [{tag}]'
            elif f'[{tag}]' not in p and tag == 'S' and '[SV]' in p:
                parts[i] = p.replace('[SV]', '[S,SV]')
            hit = True
    if not hit:
        parts.append(f'{brand} [{tag}]')
    return ', '.join(parts)

def normkey(company, street):
    s = unicodedata.normalize('NFKD', (str(company)+str(street)).lower()).encode('ascii','ignore').decode()
    return re.sub(r'[^a-z0-9]', '', s)[:40]

today = '2026-07-13'
stats = {'tagged': 0, 'new': 0, 'brands': []}
new_rows = {}  # key -> {name, company, street, psc, city, market, brands:{brand:tag}}
next_num = max([int(ws.cell(r, C_NUM).value) for r in range(hdr_row+1, ws.max_row+1)
                if str(ws.cell(r, C_NUM).value or '').strip().isdigit()] + [0]) + 1

for mf in merges:
    data = json.load(open(mf, encoding='utf-8'))
    market = re.search(r'merge_([A-Z]{2})_', mf).group(1)
    for brand, v in data.items():
        stats['brands'].append((market, brand, v['total'][:60], len(v['rows'])))
        for row in v['rows']:
            if row['match'] is not None and row['match'] in rid2xl:
                xr = rid2xl[row['match']]
                ws.cell(xr, C_BR).value = tagify(ws.cell(xr, C_BR).value, brand, row['sales'], row['service'])
                ws.cell(xr, C_VER).value = today
                ws.cell(xr, C_VER+1).value = 'Perplexity/official locator'
                stats['tagged'] += 1
            else:
                key = normkey(row['company'] or row['name'], row['street'])
                e = new_rows.setdefault(key, {'name': row['name'], 'company': row.get('group') or row['company'] or row['name'],
                    'street': row['street'], 'psc': row['psc'], 'city': row['city'], 'market': market, 'brands': {}})
                tag = 'S' if row['sales'] else 'SV'
                if row['service'] and row['sales']: tag = 'S,SV'
                prev = e['brands'].get(brand)
                if prev and 'S' in prev.split(','): tag = prev
                e['brands'][brand] = tag

for e in new_rows.values():
    r = ws.max_row + 1
    ws.cell(r, C_NUM).value = next_num; next_num += 1
    ws.cell(r, C_GRP).value = e['company']
    ws.cell(r, C_MKT).value = e['market']
    ws.cell(r, C_CITY).value = e['name']
    ws.cell(r, C_ADDR).value = f"{e['street']}, {e['psc']} {e['city']}".strip(', ')
    ws.cell(r, C_BR).value = ', '.join(f'{b} [{t}]' for b, t in sorted(e['brands'].items()))
    if C_TYPE: ws.cell(r, C_TYPE).value = 'Independent Dealer'
    ws.cell(r, C_VER).value = today
    ws.cell(r, C_VER+1).value = 'Perplexity/official locator (NEW)'
    stats['new'] += 1

# Research Status update
if 'Research Status' in wb.sheetnames:
    rs = wb['Research Status']
    idx = {(str(rs.cell(r,1).value), str(rs.cell(r,2).value)): r for r in range(2, rs.max_row+1)}
    for market, brand, total, nrows in stats['brands']:
        key = (market, brand)
        r = idx.get(key)
        if r is None:
            r = rs.max_row + 1
            rs.cell(r,1).value = market; rs.cell(r,2).value = brand
        mnum = re.search(r'(\d+)\s*(?:sales outlets|outlets with new-car sales|sales/service outlets|authorized dealers|authorized partner locations|confirmed sales/service outlets|confirmed authorized dealers|dealers listed|sales)', total)
        partial = any(k in total.lower() for k in ('enumerated','inaccessible','not stated','~','approx','firmy.cz','secondary'))
        if mnum:
            rs.cell(r,3).value = int(mnum.group(1))
            rs.cell(r,4).value = 'Official locator via Perplexity'
            rs.cell(r,5).value = today
            rs.cell(r,6).value = 'partial (enumerated)' if partial else 'verified'
        rs.cell(r,11).value = f'yes ({today})'
        rs.cell(r,12).value = f'yes — {nrows} outlets'

wb.save(WB)
print(f"applied: {stats['tagged']} existing locations tagged, {stats['new']} new dealers appended")
for m,b,t,n in stats['brands']: print(f'  {m} {b:<14} {n:>3} rows | {t}')
